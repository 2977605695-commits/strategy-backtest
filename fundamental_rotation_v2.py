"""
基本面评分 + MA5偏离 + Trail轮动 v2 · 细分赛道(小类)去重
==========================================================
v2 改动: 使用Excel股票细分赛道清单.xlsx中的"细分赛道"(小类)作为去重维度
         而非原来CSV中的大类赛道
Grid: DEV × Trail
"""
import json, os, math, csv
from collections import defaultdict

DATA_DIR = r"C:\Users\home\Desktop\strategy-backtest\data"
FUND_DIR = os.path.join(DATA_DIR, "fundamentals_70stocks")
RISK_FREE = 0.025; TRADING_DAYS = 252; INIT_CAP = 10_000_000
MA_WIN = 5; MAX_POSITIONS = 5
BUY_SLIPPAGE = 0.003; SELL_SLIPPAGE = 0.003
BUY_FEE = 0.00025; SELL_FEE = 0.00075
W_NET_MARGIN = 0.50; W_ROE = 0.37; W_REV_YOY = 0.13

def load_subsector_map():
    """Parse Excel to get code -> sub-sector mapping (fill forward blanks)"""
    import openpyxl
    wb = openpyxl.load_workbook(r"C:\Users\home\Desktop\股票细分赛道清单.xlsx")
    ws = wb[wb.sheetnames[0]]

    code_to_subsector = {}
    current_industry = ""
    current_subsector = ""

    for row in ws.iter_rows(min_row=2, values_only=True):
        industry = str(row[0]).strip() if row[0] else ""
        subsector = str(row[1]).strip() if row[1] else ""
        code = str(row[2]).strip() if row[2] else ""
        name = str(row[3]).strip() if row[3] else ""

        if industry: current_industry = industry
        if subsector: current_subsector = subsector

        if code and code.isdigit():
            code_to_subsector[code] = current_subsector

    return code_to_subsector

def calc_ma(data, w):
    ma = []
    for i in range(len(data)):
        if i < w-1: ma.append(float('nan'))
        else: ma.append(sum(data[i-w+1:i+1])/w)
    return ma

def load_fundamentals():
    fund_data = defaultdict(list)
    for fname in sorted(os.listdir(FUND_DIR)):
        if not fname.endswith('.csv'): continue
        with open(os.path.join(FUND_DIR, fname), 'r', encoding='utf-8-sig') as f:
            for row in csv.DictReader(f):
                try:
                    fund_data[row['code'].strip()].append({
                        'pub_date': row['pub_date'].strip(),
                        'report_date': row['report_date'].strip(),
                        'roe': float(row['roe']), 'net_margin': float(row['net_margin']),
                        'rev_yoy': float(row['rev_yoy']),
                    })
                except: pass
    for code in fund_data: fund_data[code].sort(key=lambda x: x['pub_date'])
    return fund_data

def get_latest(fund_data, date_str):
    latest = {}
    for code, reports in fund_data.items():
        valid = [r for r in reports if r['pub_date'] <= date_str]
        if valid: latest[code] = valid[-1]
    return latest

def compute_zscores(latest_fund):
    if len(latest_fund) < 3: return {}
    metrics = {'roe': [], 'net_margin': [], 'rev_yoy': []}
    codes = []
    for code, fund in latest_fund.items():
        codes.append(code)
        metrics['roe'].append(fund['roe'])
        metrics['net_margin'].append(fund['net_margin'])
        metrics['rev_yoy'].append(fund['rev_yoy'])
    stats = {}
    for key, vals in metrics.items():
        mu = sum(vals)/len(vals); var = sum((v-mu)**2 for v in vals)/len(vals)
        stats[key] = (mu, math.sqrt(var) if var > 0 else 1.0)
    scores = {}
    for i, code in enumerate(codes):
        z_roe = (metrics['roe'][i]-stats['roe'][0])/stats['roe'][1]
        z_nm = (metrics['net_margin'][i]-stats['net_margin'][0])/stats['net_margin'][1]
        z_ry = (metrics['rev_yoy'][i]-stats['rev_yoy'][0])/stats['rev_yoy'][1]
        scores[code] = z_nm*W_NET_MARGIN + z_roe*W_ROE + z_ry*W_REV_YOY
    return scores

def backtest_rotation(buy_thr, trail_pct, stocks_data, fund_data, subsector_map, common_dates):
    common_set = set(common_dates)
    precomputed = {}
    for code, info in stocks_data.items():
        bars = [b for b in info['bars'] if b['date'] in common_set]
        closes = [b['close'] for b in bars]
        ma5 = calc_ma(closes, MA_WIN)
        devs = []
        for i, bar in enumerate(bars):
            ma = ma5[i]
            if math.isnan(ma) or ma == 0: devs.append(float('nan'))
            else: devs.append((bar['close']-ma)/abs(ma))
        precomputed[code] = {
            'bars': bars, 'closes': closes, 'ma5': ma5, 'devs': devs,
            'subsector': subsector_map.get(code, info.get('sector', '?')),
            'name': info['name'],
        }

    per_stock_cap = INIT_CAP / MAX_POSITIONS
    holdings = {}; cash_pool = INIT_CAP; trades = []; daily_values = []
    current_scores = {}; trail_count = 0; final_count = 0

    for day_idx, date_str in enumerate(common_dates):
        # Fund update
        new_fund = get_latest(fund_data, date_str)
        if new_fund:
            new_scores = compute_zscores(new_fund)
            if new_scores: current_scores = new_scores

        # Trail check
        sell_events = []
        for code, h in list(holdings.items()):
            pc = precomputed[code]
            px = pc['bars'][day_idx]['close']
            if day_idx > h['buy_day']:
                if px > h['peak']: h['peak'] = px
                if px <= h['peak'] * (1 - trail_pct):
                    sell_px = px * (1 - SELL_SLIPPAGE)
                    gross = h['pos'] * sell_px
                    net_cash = gross - gross * SELL_FEE
                    ret = (sell_px - h['buy_px']) / h['buy_px']
                    trades.append({'code': code, 'name': pc['name'], 'ret': ret, 'exit': 'trail'})
                    trail_count += 1
                    sell_events.append((code, net_cash))

        for code, cr in sell_events:
            cash_pool += cr; del holdings[code]

        # Buy: rank by score, filter by subsector (小类)
        held_subsectors = set(h['subsector'] for h in holdings.values())
        held_codes = set(holdings.keys())
        eligible = []
        # Add subsector info from precomputed
        scored = []
        for code, score in current_scores.items():
            if code in held_codes: continue
            subsec = subsector_map.get(code, '?')
            if subsec in held_subsectors: continue
            if code not in precomputed: continue
            scored.append((code, score, subsec))

        for code, score, subsec in sorted(scored, key=lambda x: x[1], reverse=True):
            dev = precomputed[code]['devs'][day_idx]
            if not math.isnan(dev) and dev < buy_thr:
                eligible.append(code)

        while len(holdings) < MAX_POSITIONS and eligible and cash_pool >= per_stock_cap:
            code = eligible.pop(0)
            pc = precomputed[code]
            px = pc['bars'][day_idx]['close']
            buy_px = px * (1 + BUY_SLIPPAGE)
            fee = per_stock_cap * BUY_FEE
            pos = (per_stock_cap - fee) / buy_px
            holdings[code] = {
                'pos': pos, 'buy_px': buy_px, 'peak': px,
                'buy_day': day_idx, 'buy_date': date_str,
                'subsector': subsector_map.get(code, '?'),
            }
            cash_pool -= per_stock_cap
            held_subsectors.add(subsector_map.get(code, '?'))

        pv = cash_pool
        for code, h in holdings.items():
            pv += h['pos'] * precomputed[code]['bars'][day_idx]['close']
        daily_values.append({'date': date_str, 'value': pv, 'positions': len(holdings)})

    # Final
    for code, h in list(holdings.items()):
        pc = precomputed[code]
        fp = pc['bars'][-1]['close']
        sell_px = fp * (1 - SELL_SLIPPAGE)
        gross = h['pos'] * sell_px
        net_cash = gross - gross * SELL_FEE
        ret = (sell_px - h['buy_px']) / h['buy_px']
        trades.append({'code': code, 'name': pc['name'], 'ret': ret, 'exit': 'final'})
        final_count += 1
        cash_pool += net_cash; del holdings[code]

    # Metrics
    fv = daily_values[-1]['value']
    rets = []
    for i in range(1, len(daily_values)):
        p, c = daily_values[i-1]['value'], daily_values[i]['value']
        if p > 0: rets.append((c-p)/p)
    peak_v = daily_values[0]['value']; mdd = 0.0
    for dv in daily_values:
        if dv['value'] > peak_v: peak_v = dv['value']
        dd = (peak_v-dv['value'])/peak_v
        if dd > mdd: mdd = dd
    tr = (fv-INIT_CAP)/INIT_CAP
    if len(rets) > 1:
        mu = sum(rets)/len(rets); sd = (sum((r-mu)**2 for r in rets)/(len(rets)-1))**0.5
        av = sd*math.sqrt(TRADING_DAYS); ar_ = mu*TRADING_DAYS
        sh = (ar_-RISK_FREE)/av if av>0 else 0
    else: av = sh = ar_ = 0.0
    cagr = (1+tr)**(TRADING_DAYS/max(len(rets),1))-1 if tr>-1 else -1
    cm = cagr/mdd if mdd>0 else float('inf')
    wins = sum(1 for t in trades if t['ret']>0); wr = wins/len(trades) if trades else 0
    pos_days = sum(1 for dv in daily_values if dv['positions']>0)
    avg_pos = sum(dv['positions'] for dv in daily_values)/len(daily_values)

    return {
        'buy_thr': buy_thr, 'trail_pct': trail_pct,
        'tr': tr, 'ar': cagr, 'av': av, 'sh': sh, 'mdd': mdd, 'cm': cm,
        'np': len(trades), 'wr': wr, 'trail_count': trail_count, 'final_count': final_count,
        'pos_days': pos_days, 'avg_positions': avg_pos, 'fv': fv,
    }

def main():
    print("=" * 80)
    print("  基本面评分 + MA5偏离 + Trail轮动 v2 · 细分赛道(小类)去重")
    print(f"  v2: 使用Excel清单中的细分赛道, 而非大类")
    print("=" * 80)

    # Load subsector mapping
    subsector_map = load_subsector_map()
    print(f"\n[LOAD] Sub-sector map: {len(subsector_map)} codes")
    # Show unique sub-sectors
    unique_subs = sorted(set(subsector_map.values()))
    print(f"  Unique sub-sectors: {len(unique_subs)}")
    for s in unique_subs:
        codes_in = [c for c, ss in subsector_map.items() if ss == s]
        print(f"    {s}: {len(codes_in)} stocks")

    # Load fundamentals
    fund_data = load_fundamentals()
    # Load price data
    stocks = {}
    for fname in sorted(os.listdir(DATA_DIR)):
        if not fname.endswith('.json') or fname.startswith('_'): continue
        with open(os.path.join(DATA_DIR, fname), 'r', encoding='utf-8') as f:
            data = json.load(f)
        stocks[data['code']] = {'name': data['name'], 'bars': data['bars']}

    qualified = {c: s for c, s in stocks.items() if c in fund_data and len(s['bars']) >= 500}
    date_sets = [set(b['date'] for b in s['bars']) for s in qualified.values()]
    common_dates = sorted(date_sets[0].intersection(*date_sets[1:]))
    print(f"\n  {len(qualified)} stocks, {len(common_dates)} days")

    # Grid
    buy_thresholds = [-0.030, -0.035, -0.040, -0.045, -0.050, -0.055]
    trail_pcts = [0.15, 0.17, 0.20, 0.22, 0.25, 0.27, 0.30, 0.35]

    print(f"\n[GRID] {len(buy_thresholds)}x{len(trail_pcts)} = {len(buy_thresholds)*len(trail_pcts)} scenarios\n")

    all_results = []
    done = 0; total = len(buy_thresholds)*len(trail_pcts)

    for buy_thr in buy_thresholds:
        for trail_pct in trail_pcts:
            done += 1
            r = backtest_rotation(buy_thr, trail_pct, qualified, fund_data, subsector_map, common_dates)
            all_results.append(r)
            print(f"  [{done:>2d}/{total}] DEV<{buy_thr:.1%} Trail{trail_pct:.0%} | "
                  f"Sharpe={r['sh']:>7.4f} | Ret={r['tr']*100:>7.2f}% | "
                  f"DD={r['mdd']*100:>5.2f}% | Trd={r['np']:>3d} | "
                  f"Win={r['wr']*100:>5.1f}% | 持仓={r['avg_positions']:.1f}")

    # Rank
    sorted_all = sorted(all_results, key=lambda r: r['sh'], reverse=True)

    print(f"\n\n{'='*120}")
    print(f"  排名（按夏普）")
    print(f"{'='*120}")
    print(f"  {'#':<3s} {'策略':<26s} {'夏普':>7s} {'总收益':>8s} {'年化':>7s} "
          f"{'回撤':>7s} {'卡玛':>7s} {'交易':>5s} {'胜率':>6s} {'均持仓':>6s}")
    print(f"  {'-'*105}")

    for rank, r in enumerate(sorted_all, 1):
        tag = " << BEST" if rank == 1 else ""
        print(f"  {rank:<3d} DEV<{r['buy_thr']:.1%} Trail{r['trail_pct']:.0%}     "
              f"{r['sh']:>7.4f} {r['tr']*100:>7.2f}% {r['ar']*100:>6.2f}% "
              f"{r['mdd']*100:>6.2f}% {r['cm']:>7.3f} "
              f"{r['np']:>5d} {r['wr']*100:>5.1f}% {r['avg_positions']:>5.1f}{tag}")

    # Heatmap
    print(f"\n\n  --- 夏普热力图 ---")
    print(f"  {'':>12s}", end="")
    for t in trail_pcts: print(f"  Tr{t:.0%}", end="")
    print()
    for buy_thr in buy_thresholds:
        print(f"  {'DEV<'+f'{buy_thr:.1%}':<12s}", end="")
        for trail_pct in trail_pcts:
            r = [x for x in all_results if x['buy_thr']==buy_thr and x['trail_pct']==trail_pct][0]
            is_max = r['sh'] == max(x['sh'] for x in all_results)
            m = f"*{r['sh']:.3f}*" if is_max else f" {r['sh']:.3f} "
            print(f" {m}", end="")
        print()

    # Comparison with v1
    print(f"\n\n  --- v1(大类) vs v2(小类) 对比 ---")
    # Best from v1: DEV<-3.5% Trail35% -> Sharpe 2.3720
    # Find same params in v2
    for r in sorted_all:
        if abs(r['buy_thr'] - (-0.035)) < 0.001 and abs(r['trail_pct'] - 0.35) < 0.001:
            print(f"  DEV<-3.5% Trail35%: v1=2.3720  v2={r['sh']:.4f}  diff={r['sh']-2.3720:+.4f}")
            break

    print(f"\n  Done!")

if __name__ == '__main__':
    main()
